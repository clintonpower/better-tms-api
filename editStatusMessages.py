import csv                      # type: ignore
import openpyxl                 # type: ignore
import requests                 # type: ignore
from datetime import datetime   # type: ignore
from urllib.parse import quote  # type: ignore
from bs4 import BeautifulSoup   # type: ignore
import re                       # type: ignore

def load_config(config_sheet):
    config = {}
    # Read configuration from columns A (key) and B (value)
    for row in config_sheet.iter_rows(min_row=1, values_only=True):
        if row[0] and row[1]:
            config[row[0].strip()] = str(row[1]).strip()
    return config

def load_mapping(csv_filename):
    mapping = {}
    try:
        with open(csv_filename, newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                tid = row.get("transport_id")
                toid = row.get("transport_order_id")
                if tid and toid:
                    mapping[tid.strip()] = toid.strip()
        print(f"Loaded mapping for {len(mapping)} transport IDs from {csv_filename}.")
    except Exception as e:
        print(f"Error reading mapping CSV file: {e}")
    return mapping

def get_csrf_token(session, primary_server):
    url = f"https://{primary_server}.mercurygate.net/MercuryGate/transport/addMessage.jsp?norefresh=&messageCode=AF"
    try:
        resp = session.get(url, timeout=10)
        if resp.status_code == 200:
            soup = BeautifulSoup(resp.text, "html.parser")
            meta = soup.find("meta", {"name": "_csrf"})
            if meta and meta.has_attr("content"):
                token = meta["content"]
                print("Extracted CSRF token from meta tag:", token)
                return token
            else:
                print("CSRF meta tag not found on page.")
                return ""
        else:
            print("Failed to fetch CSRF token page; status code:", resp.status_code)
            return ""
    except Exception as e:
        print("Error fetching CSRF token:", e)
        return ""

def format_transport_order_id(mapping_value, suffix):
    """
    If mapping value doesn't contain a comma, assume it is the base value and append the suffix.
    """
    if "," not in mapping_value:
        formatted = f"({mapping_value}{suffix})"
    else:
        if not mapping_value.startswith("("):
            formatted = f"({mapping_value})"
        else:
            formatted = mapping_value
    return formatted

def format_sidEvent(event_oid, event_suffix):
    return f"({event_oid},{event_suffix})"

def parse_pickup_datetime(value):
    """
    Parses the pickup date/time value.
    Expected examples:
      "3/16/2024  7:00:00 AM"  => Date: "03/16/2024", Time: "07:00 AM"
      "3/19/24 12:00"         => Date: "03/19/2024", Time: "12:00 PM" (if appropriate)
    If the cell is a datetime object, formats it accordingly.
    """
    if isinstance(value, datetime):
        return value.strftime("%m/%d/%Y"), value.strftime("%I:%M %p")
    else:
        try:
            # Clean up extra spaces.
            v = re.sub(r'\s+', ' ', str(value).strip())
            # Try parsing with full year and seconds (e.g. "3/16/2024 7:00:00 AM")
            parsed = datetime.strptime(v, "%m/%d/%Y %I:%M:%S %p")
            return parsed.strftime("%m/%d/%Y"), parsed.strftime("%I:%M %p")
        except Exception:
            try:
                # Try parsing with two-digit year and no seconds (e.g. "3/19/24 12:00")
                parsed = datetime.strptime(v, "%m/%d/%y %H:%M")
                return parsed.strftime("%m/%d/%Y"), parsed.strftime("%I:%M %p")
            except Exception as e:
                print("Failed to parse pickup date/time:", value, "Error:", e)
                return str(value).strip(), ""

def prime_session(session, primary_server):
    url = f"https://{primary_server}.mercurygate.net/MercuryGate/transport/addMessage.jsp?norefresh=&messageCode=AF"
    try:
        resp = session.get(url, timeout=10)
        print("Priming GET status:", resp.status_code)
    except Exception as e:
        print("Error during priming GET:", e)

def process_excel_and_post(excel_path, csv_mapping_path):
    wb = openpyxl.load_workbook(excel_path)
    config_sheet = wb["config"]
    config = load_config(config_sheet)
    
    required_vars = [
        "PRIMARY_SERVER",
        "AUTH_COOKIE",
        "ENTERPRISE_OID",
        "EVENT_SUFFIX",
        "STATUS_MESSAGE",
        "TRANSPORT_ORDER_SUFFIX",
        "SCAC"
    ]
    for var in required_vars:
        if var not in config:
            raise ValueError(f"Missing required config variable: {var}")
    
    primary_server = config["PRIMARY_SERVER"]
    auth_cookie = config["AUTH_COOKIE"]
    
    mapping = load_mapping(csv_mapping_path)
    if not mapping:
        print("No mapping data loaded. Exiting.")
        return
    
    session = requests.Session()
    session.headers.update({
        "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
        "accept-language": "en-GB,en-US;q=0.9,en;q=0.8,de;q=0.7",
        "cache-control": "max-age=0",
        "content-type": "application/x-www-form-urlencoded",
        "dnt": "1",
        "origin": f"https://{primary_server}.mercurygate.net",
        "priority": "u=0, i",
        "referer": f"https://{primary_server}.mercurygate.net/MercuryGate/transport/addMessage_process.jsp",
        "sec-ch-ua": "\"Chromium\";v=\"134\", \"Not:A-Brand\";v=\"24\", \"Google Chrome\";v=\"134\"",
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": "\"macOS\"",
        "sec-fetch-dest": "document",
        "sec-fetch-mode": "navigate",
        "sec-fetch-site": "same-origin",
        "sec-fetch-user": "?1",
        "upgrade-insecure-requests": "1",
        "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36",
        "cookie": auth_cookie
    })
    
    # Prime the session.
    prime_session(session, primary_server)
    
    post_url = f"https://{primary_server}.mercurygate.net/MercuryGate/transport/addMessage_process.jsp"
    
    lookup_sheet = wb["lookup"]
    header_row = [cell.value for cell in lookup_sheet[1]]
    if "Status" in header_row:
        status_col = header_row.index("Status") + 1
    else:
        status_col = len(header_row) + 1
        lookup_sheet.cell(row=1, column=status_col, value="Status")
    
    for idx, row in enumerate(lookup_sheet.iter_rows(min_row=2, values_only=True), start=2):
        row_data = dict(zip(header_row, row))
        if not row_data.get("Shipping Order"):
            print(f"Empty 'Shipping Order' at row {idx}. Stopping processing.")
            break

        # Get transport ID from "SO Oid"
        transport_id = row_data.get("SO Oid")
        if not transport_id:
            msg = "Missing SO Oid (transport_id)"
            print(f"Row {idx}: {msg}")
            lookup_sheet.cell(row=idx, column=status_col, value=msg)
            continue
        transport_id = str(transport_id).strip()
        
        base_mapping = mapping.get(transport_id)
        if not base_mapping:
            msg = f"Mapping not found for transport_id {transport_id}"
            print(f"Row {idx}: {msg}")
            lookup_sheet.cell(row=idx, column=status_col, value=msg)
            continue
        transport_order_id = format_transport_order_id(base_mapping, config["TRANSPORT_ORDER_SUFFIX"])
        
        # Get Event Oid and build sidEvent.
        event_oid = row_data.get("Event Oid")
        if not event_oid:
            msg = "Missing Event Oid"
            print(f"Row {idx}: {msg}")
            lookup_sheet.cell(row=idx, column=status_col, value=msg)
            continue
        event_oid = str(event_oid).strip()
        sidEvent = format_sidEvent(event_oid, config["EVENT_SUFFIX"])
        
        # Parse the pickup date and time.
        pickup_date_raw = row_data.get("Pickup Date")
        date_str, time_str = parse_pickup_datetime(pickup_date_raw)
        
        # For PRO, use the value from the "Shipping Order" column.
        pro_value = row_data.get("Shipping Order")
        if pro_value:
            pro_value = str(pro_value).strip()
        else:
            pro_value = ""
        scac_value = config["SCAC"]
        
        post_payload = {
            "norefresh": "",
            "bRefresh": "false",
            "oidEnterprise": config["ENTERPRISE_OID"],
            "bShowReferences": "true",
            "sidTransportOrder": transport_order_id,
            "sidEvent": sidEvent,
            "requireApproval": "false",
            "changeRequestType": "",
            "changeRequestOwnerOid": "",
            "sEvent": "",
            "sOrigApptComment": "",
            "SCAC": scac_value,
            "PRO": pro_value,
            "sType": "AF",
            "dateDate1": "",
            "dateTime1": "",
            "dateDate2": date_str,
            "dateTime2": "12:00 PM", # hardcoded
            #"dateTime2": time_str, # dynamic    
            "sLateReasonCode": "",
            "sidReferenceType1": "(100106,3250,0)", # TODO: Check if this can be blank
            "sReference1": ""
        }
        
        try:
            response = session.post(post_url, data=post_payload, timeout=10)
            #print("=== POST Debug (Row", idx, ") ===")
            #print("POST URL:", post_url)
            #print("POST payload:", post_payload)
            #print("Response Status Code:", response.status_code)
            #print("Response (first 300 chars):", response.text[:300])
            if response.status_code == 200:
                result_text = "200 OK"
            else:
                result_text = f"{response.status_code}: {response.text[:100]}"
            print(f"Row {idx} -> transport_order_id: {transport_order_id} | {result_text}")
            lookup_sheet.cell(row=idx, column=status_col, value=result_text)
        except Exception as e:
            err_msg = f"Error: {str(e)}"
            print(f"Row {idx} -> {err_msg}")
            lookup_sheet.cell(row=idx, column=status_col, value=err_msg)
    
    output_path = "./OrdersToBeUpdated_tmp_updated.xlsx"
    wb.save(output_path)
    print(f"Processing complete. Results saved to {output_path}")

if __name__ == "__main__":
    excel_file_path = "./OrdersToBeUpdated_statusmessages.xlsx"
    csv_mapping_path = "./All_SO_Data.csv"
    process_excel_and_post(excel_file_path, csv_mapping_path)
