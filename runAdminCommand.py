import openpyxl                 # type: ignore
import requests                 # type: ignore
from urllib.parse import quote  # type: ignore
from bs4 import BeautifulSoup   # type: ignore
import re                       # type: ignore

def load_config(config_sheet):
    config = {}
    for row in config_sheet.iter_rows(min_row=1, values_only=True):
        if row[0] and row[1]:
            config[row[0].strip()] = str(row[1]).strip()
    return config

def prime_session(session, primary_server):
    url = f"https://{primary_server}.mercurygate.net/MercuryGate/util/adminConsole.jsp"
    try:
        resp = session.get(url, timeout=10)
        print("Priming GET status:", resp.status_code)
    except Exception as e:
        print("Error during priming GET:", e)

def parse_response_message(html_text):
    soup = BeautifulSoup(html_text, "html.parser")
    script_tags = soup.find_all("script")
    for script in script_tags:
        if "displayWindow('Results', message);" in script.text:
            match = re.search(r"var message = '(.*?)';", script.text, re.DOTALL)
            if match:
                return match.group(1).replace("\\n", "\n")
    return "No message found"

def run_commands(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    config_sheet = wb["config"]
    lookup_sheet = wb["lookup"]

    config = load_config(config_sheet)
    required_keys = ["PRIMARY_SERVER", "AUTH_COOKIE", "ENTERPRISE"]
    for key in required_keys:
        if key not in config:
            raise ValueError(f"Missing required config key: {key}")

    primary_server = config["PRIMARY_SERVER"]
    auth_cookie = config["AUTH_COOKIE"]
    enterprise = config["ENTERPRISE"]
    sid_enterprise = quote(f"({enterprise},3640,0)")

    session = requests.Session()
    session.headers.update({
        "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "content-type": "application/x-www-form-urlencoded",
        "user-agent": "Mozilla/5.0",
        "origin": f"https://{primary_server}.mercurygate.net",
        "referer": f"https://{primary_server}.mercurygate.net/MercuryGate/util/adminConsole.jsp",
        "cookie": auth_cookie
    })

    prime_session(session, primary_server)

    # Setup result column
    header_row = [cell.value for cell in lookup_sheet[1]]
    if "Result" in header_row:
        result_col = header_row.index("Result") + 1
    else:
        result_col = len(header_row) + 1
        lookup_sheet.cell(row=1, column=result_col, value="Result")

    for idx, row in enumerate(lookup_sheet.iter_rows(min_row=2, values_only=True), start=2):
        command = row[0]
        if not command:
            print(f"Stopping at empty row {idx}")
            break

        post_url = f"https://{primary_server}.mercurygate.net/MercuryGate/util/adminConsole.jsp?sidEnterprise={sid_enterprise}&"
        post_data = {"sCommandList": command}

        try:
            response = session.post(post_url, data=post_data, timeout=10)
            if response.status_code == 200:
                message = parse_response_message(response.text)
            else:
                message = f"HTTP {response.status_code}: {response.text[:100]}"
        except Exception as e:
            message = f"Error: {str(e)}"

        print(f"Row {idx}: Command '{command}' -> {message}")
        lookup_sheet.cell(row=idx, column=result_col, value=message)

    output_path = "./runAdminCommand_updated.xlsx"
    wb.save(output_path)
    print(f"Processing complete. Results saved to {output_path}")

if __name__ == "__main__":
    excel_file_path = "./runAdminCommand.xlsx"  # Update this path as needed
    run_commands(excel_file_path)
