import openpyxl                                                     # type: ignore
import requests                                                     # type: ignore
from urllib.parse import quote_plus                                 # type: ignore
from collections import defaultdict                                 # type: ignore
from concurrent.futures import ThreadPoolExecutor, as_completed     # type: ignore

# --- Helpers ---

def load_config(sheet):
    """Load key-value config data from the 'config' sheet."""
    config = {}
    for row in sheet.iter_rows(min_row=1, values_only=True):
        if row[0] and row[1]:
            config[row[0].strip()] = str(row[1]).strip()
    return config

def encode_sid(enterprise_oid):
    """Encode the sidEnterprise format used in TMS URLs."""
    return quote_plus(f"({enterprise_oid},3640,0)")

def group_settings_by_page(sheet):
    """Group setting changes by page name for batch POST requests."""
    grouped = defaultdict(dict)
    header = [cell.value for cell in sheet[1]]
    page_idx = header.index("page")
    setting_idx = header.index("setting")
    value_idx = header.index("value")

    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Stop if any required field is missing (first empty row)
        if not row[page_idx] or not row[setting_idx] or not row[value_idx]:
            break
        page = str(row[page_idx]).strip()
        setting = str(row[setting_idx]).strip()
        value = str(row[value_idx]).strip()
        grouped[page][setting] = value
    return grouped

def ensure_status_column(sheet):
    """Add a Status column to the sheet if missing and return its index."""
    header = [cell.value for cell in sheet[1]]
    if "Status" in header:
        return header.index("Status") + 1
    col = len(header) + 1
    sheet.cell(row=1, column=col, value="Status")
    return col

def prime_session(session, primary_server):
    """Make a priming GET request to warm up the session."""
    url = f"https://{primary_server}.mercurygate.net/MercuryGate/enterprise/editEnterpriseSysConMisc.jsp"
    try:
        resp = session.get(url, timeout=10)
        print("Priming GET status:", resp.status_code)
    except Exception as e:
        print("Error during priming GET:", e)

def post_settings(page, settings, sidEnterprise, config, primary_server):
    """Send a single POST request for one settings page with all its settings."""
    url = f"https://{primary_server}.mercurygate.net/MercuryGate/enterprise/{page}"
    referer_url = f"https://{primary_server}.mercurygate.net/MercuryGate/enterprise/{page.replace('_process', '')}"

    # Build raw form body
    form_items = [f"sidEnterprise={sidEnterprise}"] + [f"{k}={v}" for k, v in settings.items()]
    body_str = "&".join(form_items)

    headers = {
        "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
        "accept-language": "en-GB,en-US;q=0.9,en;q=0.8,de;q=0.7",
        "cache-control": "max-age=0",
        "content-type": "application/x-www-form-urlencoded",
        "priority": "u=0, i",
        "sec-ch-ua": '"Chromium";v="134", "Not:A-Brand";v="24", "Google Chrome";v="134"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"macOS"',
        "sec-fetch-dest": "iframe",
        "sec-fetch-mode": "navigate",
        "sec-fetch-site": "same-origin",
        "sec-fetch-user": "?1",
        "upgrade-insecure-requests": "1",
        "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36",
        "referer": referer_url,
        "origin": f"https://{primary_server}.mercurygate.net",
        "cookie": config["AUTH_COOKIE"]
    }

    try:
        resp = requests.post(url, data=body_str, headers=headers, timeout=15)
        if resp.status_code == 200:
            return "200 - OK"
        else:
            return f"{resp.status_code} - {resp.text[:100]}"
    except Exception as e:
        return f"Error: {str(e)}"

# --- Main Processing ---
def process_sysconfigs(excel_path, max_workers=10):
    """Main entry point for processing sysconfig updates from Excel file."""
    wb = openpyxl.load_workbook(excel_path)
    config = load_config(wb["config"])
    lookup_sheet = wb["lookup"]

    required_keys = ["PRIMARY_SERVER", "AUTH_COOKIE", "ENTERPRISE"]
    for k in required_keys:
        if k not in config:
            raise ValueError(f"Missing required config key: {k}")

    primary_server = config["PRIMARY_SERVER"]
    sidEnterprise = quote_plus(f"({config['ENTERPRISE']},3640,0)")

    # Create a session for priming
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0",
        "Cookie": config["AUTH_COOKIE"]
    })
    prime_session(session, primary_server)

    # Group settings per page and prepare for batch POSTing
    grouped = group_settings_by_page(lookup_sheet)
    status_col = ensure_status_column(lookup_sheet)

    futures = {}
    results = {}
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        for page, settings in grouped.items():
            futures[executor.submit(post_settings, page, settings, sidEnterprise, config, primary_server)] = page

        for future in as_completed(futures):
            page = futures[future]
            results[page] = future.result()

    # Record status back into Excel
    for row in range(2, lookup_sheet.max_row + 1):
        page_cell = lookup_sheet.cell(row=row, column=1).value
        if not page_cell:
            break
        page = str(page_cell).strip()
        status = results.get(page, "Not attempted")
        lookup_sheet.cell(row=row, column=status_col, value=status)

    # Save results
    output_path = excel_path.replace(".xlsx", "_updated.xlsx")
    wb.save(output_path)
    print(f"Finished. Results written to {output_path}")

if __name__ == "__main__":
    process_sysconfigs("./SysConfigUpdates.xlsx", max_workers=10)
