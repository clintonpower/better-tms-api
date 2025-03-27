import csv                                                          # type: ignore
import openpyxl                                                     # type: ignore
import requests                                                     # type: ignore
from datetime import datetime                                       # type: ignore
from urllib.parse import quote                                      # type: ignore
from concurrent.futures import ThreadPoolExecutor, as_completed     # type: ignore

def load_config(config_sheet):
    config = {}
    # Read configuration from columns A (key) and B (value)
    for row in config_sheet.iter_rows(min_row=1, values_only=True):
        if row[0] and row[1]:
            config[row[0].strip()] = str(row[1]).strip()
    return config

def load_mapping(csv_filename):
    mapping = {}
    try:
        with open(csv_filename, newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                tid = row.get("transport_id")
                toid = row.get("transport_order_id")
                if tid and toid:
                    mapping[tid.strip()] = toid.strip()
        print(f"Loaded mapping for {len(mapping)} transport IDs from {csv_filename}.")
    except Exception as e:
        print(f"Error reading mapping CSV file: {e}")
    return mapping

def format_transport_order_id(mapping_value, suffix):
    """
    If mapping value doesn't contain a comma, assume it is the base value and append the suffix.
    """
    if "," not in mapping_value:
        formatted = f"({mapping_value}{suffix})"
    else:
        if not mapping_value.startswith("("):
            formatted = f"({mapping_value})"
        else:
            formatted = mapping_value
    return formatted

def prime_session(session, primary_server):
    url = f"https://{primary_server}.mercurygate.net/MercuryGate/pricesheets/editPriceSheet_process.jsp"
    try:
        resp = session.get(url, timeout=10)
        print("Priming GET status:", resp.status_code)
    except Exception as e:
        print("Error during priming GET:", e)

def process_row(row_data, config, mapping, session, primary_server):
    """
    Process a single row of the Excel lookup data.
    Returns a string with the format:
      "SO {pri_ref} OK" or "SO {pri_ref} Error: <error message>"
    """
    # Ensure required fields are present:
    # "OTM_COST" for the new cost,
    # "pri_ref" as the SO number (for PRO),
    # "pricesheet_is" as oidPriceSheet,
    # "transport_order_id" if no CSV mapping is provided.
    try:
        so_number = str(row_data.get("pri_ref", "")).strip()
        otm_cost = row_data.get("OTM_COST")
        if otm_cost is None or str(otm_cost).strip() == "":
            return f"SO {so_number} Error: Missing OTM_COST"
        new_cost = str(otm_cost).strip()
        oidPriceSheet = str(row_data.get("pricesheet_is", "")).strip()
        if not oidPriceSheet:
            return f"SO {so_number} Error: Missing pricesheet_is"
        
        # Use CSV mapping if available; if not, use Excel column "transport_order_id".
        transport_id = str(row_data.get("transport_id", "")).strip()
        if transport_id and mapping:
            base_mapping = mapping.get(transport_id)
            if not base_mapping:
                return f"SO {so_number} Error: Mapping not found for transport_id {transport_id}"
            transport_order_id = format_transport_order_id(base_mapping, config["TRANSPORT_ORDER_SUFFIX"])
        else:
            # Fall back to using Excel column "transport_order_id"
            transport_order_id = str(row_data.get("transport_order_id", "")).strip()
            if transport_order_id and "," not in transport_order_id:
                transport_order_id = f"({transport_order_id}{config['TRANSPORT_ORDER_SUFFIX']})"
            elif transport_order_id and not transport_order_id.startswith("("):
                transport_order_id = f"({transport_order_id})"
        
        # Hardcode values that are not needed
        post_payload = {
            "sSheetType": "Cost",
            "listOwnerOids": "4554639789", #TODO: Remove hardcoded Stellantis enterprise
            "sReturnURL": "/MercuryGate/transport/editTransportOrig.jsp?sidTransport=(4554639789,3300,0)",
            "sPostProcessURL": "",
            "oidPriceSheet": oidPriceSheet,
            "bGLWaiver": "false",
            "isVendor": "false",
            "sidCarrier": "(4533774089,3840,0)",
            "sCarrierMode": "TL",
            "sCarrierService": "Standard",
            "fCarrierServiceDays": "",
            "oidContract": "",
            "sCurrencyCode": "EUR",
            "oidCarrierLocation": "-1",
            "CostChargeModel": "NORMALIZED_MANUAL",
            "CostCharge1Type": "ITEM",
            "CostCharge1Desc": "Total Line Haul",
            "CostCharge1EDICode": "",
            "CostCharge1Rate": new_cost,
            "CostCharge1RQ": "FR",
            "CostCharge2Type": "DISCOUNT",
            "CostCharge2Desc": "Discount",
            "CostCharge2EDICode": "DSC",
            "CostCharge2Rate": "",
            "CostCharge2RQ": "FR",
            "CostCharge3Type": "ACCESSORIAL_FUEL",
            "CostCharge3Desc": "Fuel Surcharge",
            "CostCharge3EDICode": "FUE",
            "CostCharge3Rate": "",
            "CostCharge3RQ": "FR",
            "CostCharge4Type": "ACCESSORIAL",
            "CostCharge4Desc": "",
            "CostCharge4EDICode": "LFA",
            "CostCharge4Rate": "",
            "CostCharge4RQ": "FR",
            "CostCharge5Type": "ACCESSORIAL",
            "CostCharge5Desc": "",
            "CostCharge5EDICode": "LFA",
            "CostCharge5Rate": "",
            "CostCharge5RQ": "FR",
            "CostCharge6Type": "ACCESSORIAL",
            "CostCharge6Desc": "",
            "CostCharge6EDICode": "LFA",
            "CostCharge6Rate": "",
            "CostCharge6RQ": "FR",
            "CostCharge7Type": "ACCESSORIAL_PERCENTAGE_TOTAL",
            "CostCharge7Desc": "",
            "CostCharge7EDICode": "TAX:PST",
            "CostCharge7Rate": "",
            "CostCharge7RQ": "PCT",
            "CostNumCharges": "7",
            "sCommentsPS": "",
            "fDistance": "",
            "dateDate1": "",
            "dateTime1": "",
            "dateDate2": "",
            "dateTime2": ""
        }
        
        # Send the POST request.
        post_url = f"https://{primary_server}.mercurygate.net/MercuryGate/pricesheets/editPriceSheet_process.jsp"
        resp = session.post(post_url, data=post_payload, timeout=10)
        if resp.status_code == 200:
            return f"SO {so_number} OK"
        else:
            return f"SO {so_number} Error: {resp.status_code} {resp.text[:100]}"
    except Exception as e:
        return f"SO {so_number} Error: {str(e)}"

def process_pricesheets_concurrent(excel_path, csv_mapping_path, maxWorkers=10):
    wb = openpyxl.load_workbook(excel_path)
    config_sheet = wb["config"]
    config = load_config(config_sheet)
    
    required_vars = [
        "PRIMARY_SERVER",
        "AUTH_COOKIE",
        "ENTERPRISE_OID",
        "EVENT_SUFFIX",
        "STATUS_MESSAGE",
        "TRANSPORT_ORDER_SUFFIX",
        "SCAC"
    ]
    for var in required_vars:
        if var not in config:
            raise ValueError(f"Missing required config variable: {var}")
    
    primary_server = config["PRIMARY_SERVER"]
    auth_cookie = config["AUTH_COOKIE"]
    
    mapping = {}
    if csv_mapping_path:
        mapping = load_mapping(csv_mapping_path)
    
    lookup_sheet = wb["lookup"]
    header_row = [cell.value for cell in lookup_sheet[1]]
    # Stop processing at the first blank row in column "pri_ref"
    rows = []
    for row in lookup_sheet.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(header_row, row))
        if not row_dict.get("pri_ref"):
            break
        rows.append(row_dict)
    
    session = requests.Session()
    session.headers.update({
        "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
        "accept-language": "en-GB,en-US;q=0.9,en;q=0.8,de;q=0.7",
        "cache-control": "max-age=0",
        "content-type": "application/x-www-form-urlencoded",
        "dnt": "1",
        "origin": f"https://{primary_server}.mercurygate.net",
        "priority": "u=0, i",
        "referer": f"https://{primary_server}.mercurygate.net/MercuryGate/pricesheets/editPriceSheet_process.jsp",
        "sec-ch-ua": "\"Chromium\";v=\"134\", \"Not:A-Brand\";v=\"24\", \"Google Chrome\";v=\"134\"",
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": "\"macOS\"",
        "sec-fetch-dest": "document",
        "sec-fetch-mode": "navigate",
        "sec-fetch-site": "same-origin",
        "sec-fetch-user": "?1",
        "upgrade-insecure-requests": "1",
        "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36",
        "cookie": auth_cookie
    })
    
    # Prime the session.
    prime_session(session, primary_server)
    
    results = []
    with ThreadPoolExecutor(max_workers=maxWorkers) as executor:
        future_to_row = {executor.submit(process_row, row, config, mapping, session, primary_server): row for row in rows}
        for future in as_completed(future_to_row):
            res = future.result()
            print(res)
            results.append(res)
    
    print("Processing complete.")
    
if __name__ == "__main__":
    excel_file_path = "./OrdersToBeUpdated_pricesheet.xlsx"
    csv_mapping_path = None  
    maxWorkers = 20
    process_pricesheets_concurrent(excel_file_path, csv_mapping_path, maxWorkers)
